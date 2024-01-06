import csv
import os
from PIL import Image
from openpyxl.reader.excel import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import subprocess
from reportlab.platypus import SimpleDocTemplate, Table,TableStyle, PageBreak
from Src.constants.constants import LIBREOFFICE_PATH

class Converter:
    def convert_image_to_pdf(self, input_path, output_path):
        # A4 size
        pdf_canvas = canvas.Canvas(output_path, pagesize=(595, 842))
        img = Image.open(input_path)
        img_width, img_height = img.size
        scale_factor = min(595 / img_width, 842 / img_height)
        x_offset = (595 - img_width * scale_factor) / 2
        y_offset = (842 - img_height * scale_factor) / 2
        # Draw the image on the PDF
        pdf_canvas.drawInlineImage(input_path, x_offset, y_offset, width=img_width * scale_factor,
                               height=img_height * scale_factor)
        # Save the PDF
        pdf_canvas.save()


    # Function to convert document types to PDF
    def convert_doc_to_pdf(self, input_path, output_path):
        try:
        # Using libreoffice to convert the document to PDF
            subprocess_args = [
            'libreoffice',
            '--headless',
            '--convert-to', 'pdf:writer_pdf_Export',
            '--outdir', os.path.dirname(output_path),
            '--writer_pdf_Export_PageSize', 'A4',
            input_path
            ]
            subprocess.run(subprocess_args, check=True)
            return True
        except subprocess.CalledProcessError as e:
            self.logger.error(f"Conversion failed: {e}")
            return False

    def convert_ppt_to_pdf(self, input_path, output_path):
        try:
            # Using libre office for converting ppt to PDF
            env = os.environ.copy()
            env['PDFA1B_OUTDIR'] = os.path.dirname(output_path)
            subprocess_args = [os.path.join(LIBREOFFICE_PATH, 'libreoffice'), '--headless', '--convert-to',
                               'pdf:writer_pdf_Export', input_path]
            subprocess.run(subprocess_args, check=True, env=env)
            return True
        except subprocess.CalledProcessError as e:
            print(f"Conversion failed: {e}")
            return False

    # Function to convert CSV to PDF
    def convert_csv_to_pdf(self, input_path, output_path):
        # Reading the CSV file
        data = []
        with open(input_path, 'r') as csvfile:
            csvreader = csv.reader(csvfile)
            for row in csvreader:
                data.append(row)

        # Create a PDF template using reportlab
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        elements = []

        font_name = 'Helvetica-Bold'
        font_size = 8

        columns_per_table = 5

        num_columns = len(data[0])

        for start_col in range(0, num_columns, columns_per_table):
            end_col = start_col + columns_per_table
            table_data = [row[start_col:end_col] for row in data]

            # Generate table with data for PDF
            table = Table(table_data)

            col_widths = [1.5] * len(table_data[0])
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), (0.9, 0.9, 0.9)),
                ('TEXTCOLOR', (0, 0), (-1, 0), (0, 0, 0)),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), (0.85, 0.85, 0.85)),
                ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
                ('FONTSIZE', (0, 0), (-1, -1), font_size),
            ])
            table.setStyle(style)
            elements.append(table)

            if end_col < num_columns:
                elements.append(PageBreak())

        doc.build(elements)

    # Function to convert Excel to PDF
    def convert_excel_to_pdf(self, input_path, output_path):
        # Create a landscape PDF using the reportlab library
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        elements = []
        workbook = load_workbook(input_path)

        # Font size and style
        font_name = 'Helvetica-Bold'
        font_size = 12

        # Limiting colunm in pdf
        columns_per_table = 5

        # Reading Sheets from Excel
        for sheet in workbook.sheetnames:
            data = []
            for row in workbook[sheet].iter_rows(values_only=True):
                data.append(row)

            num_columns = len(data[0])

            for start_col in range(0, num_columns, columns_per_table):
                end_col = start_col + columns_per_table
                table_data = [row[start_col:end_col] for row in data]

                # Generate table with data for PDF
                table = Table(table_data)

                col_widths = [1.5] * len(table_data[0])
                style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), (0.8, 0.8, 0.8)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), (1, 1, 1)),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), font_name),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), (0.95, 0.95, 0.95)),
                    ('FONTSIZE', (0, 0), (-1, -1), font_size),
                ])
                table.setStyle(style)
                elements.append(table)
                if end_col < num_columns:
                    elements.append(PageBreak())
        doc.build(elements)
